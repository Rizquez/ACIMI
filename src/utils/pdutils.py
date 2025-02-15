# -------------------------------------------------------------------------------------------------------------------------------------------------
# LIBRERIAS (EXTERNAS)
# -------------------------------------------------------------------------------------------------------------------------------------------------
import openpyxl.utils
import openpyxl.worksheet.table
import openpyxl.worksheet
import pandas as pd
from io import BytesIO
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from werkzeug.datastructures import FileStorage
# -------------------------------------------------------------------------------------------------------------------------------------------------

# LIBRERIAS (INTERNAS)
# -------------------------------------------------------------------------------------------------------------------------------------------------
from src.support import Templates
from src.support import STYLE_EXCEL
from src.utils._xlsx import ProcessExcel
# -------------------------------------------------------------------------------------------------------------------------------------------------

# REFERENCIA AL FICHERO LOG (INFO/WARNING/ERROR)
# -------------------------------------------------------------------------------------------------------------------------------------------------
# Se referencian aqui!
# -------------------------------------------------------------------------------------------------------------------------------------------------

# # CREACION DE LA(S) CLASE(S) / FUNCIONES GENERALES
# -------------------------------------------------------------------------------------------------------------------------------------------------

class PdUtils(pd.DataFrame):
    """
    Descripcion
    -----------
    Extiende de la clase `pandas.DataFrame`, reuniendo internamente un conjunto de metodos especializados para realizar operaciones 
    avanzadas y personalizadas con DataFrames, asi como gestionar operaciones de lectura y escritura en documentos de `Excel` utilizando 
    `pandas` y `openpyxl`.

    Detalles
    --------
    - El objetivo es facilitar tareas comunes como la insercion, modificacion y eliminacion de datos de manera eficiente.
    - La informacion detallada acerca del alcance de aplicacion de cada metodo, se encuentra internamente en su respectiva documentacion.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    @property
    def _constructor(self) -> type['PdUtils']:
        """
        Summary
        -------
        Este metodo es una las formas recomendadas para lograr que las operaciones internas de `pandas` devuelvan una instancia de nuestra 
        subclase, sobreescribiendo la propiedad `_constructor`. Al hacerlo le indicamos a `pandas` que clase utilizar para construir nuevos 
        objetos cuando se realicen operaciones sobre un DataFrame.

        Details
        -------
        - La documentacion oficial de `pandas` sobre extension de objetos menciona el uso de `_constructor` para mantener el tipo de la subclase.
        """
        return PdUtils

    @staticmethod
    def template_excel(tipo: str) -> 'PdUtils':
        """
        Descripcion
        -----------
        Genera un DataFrame a partir de uno datos bases segun el tipo de parametro indicado.

        Parametros
        ----------
        tipo
            Tipo calculo bajo el cual se creara el DataFrame, se espera como valor: `terciario` o `residencial`.

        Retorna
        -------
        - DataFrame de con los datos base segun el tipo de calculo.
        """
        if tipo.lower() == 'terciario':
            dct_data = {
                Templates.Terciario.ESPACIO: [Templates.VALUE_EMPTY],
                Templates.Terciario.PLANTA: [Templates.VALUE_EMPTY],
                Templates.Terciario.POTENCIA_REFRIGERACION_SUPERFICIE: [Templates.ZERO],
                Templates.Terciario.POTENCIA_TOTAL_REFRIGERACION: [Templates.ZERO],
                Templates.Terciario.POTENCIA_CALEFACCION_SUPERFICIE: [Templates.ZERO],
                Templates.Terciario.POTENCIA_TOTAL_CALEFACCION: [Templates.ZERO],
                Templates.Terciario.SUPERFICIE: [Templates.ZERO]
            }

        elif tipo.lower() == 'residencial':
            dct_data = {
                Templates.Residencial.ESCALERA: [Templates.VALUE_EMPTY],
                Templates.Residencial.PLANTA: [Templates.VALUE_EMPTY],
                Templates.Residencial.VIVIENDA: [Templates.VALUE_EMPTY],
                Templates.Residencial.SUPERFICIE: [Templates.ZERO],
                Templates.Residencial.LONGITUD_H_PLANTA: [Templates.ZERO],
                Templates.Residencial.LONGITUD_V_MONTANTE: [Templates.ZERO],
                Templates.Residencial.LONGITUD_H_CUBIERTA: [Templates.ZERO]
            }

        else:
            raise ValueError(f"El tipo de calculo {tipo} no se encuentra contemplado por el metodo `template_xlsx`")
        
        return PdUtils(dct_data)
    
    def save_excel(self, table_name: str, sheet_name: str, mode: str, float_format: str) -> BytesIO:
        """
        Descripcion
        -----------
        Crea o actualiza una tabla en un archivo `Excel` a partir de un `DataFrame`.

        Detalles
        --------
        - La tabla abarca desde la celda `A1` hasta el tamaño dinamico del `DataFrame`.
        - Ajusta automaticamente el ancho de las columnas y alinea las celdas.

        Parametros
        ----------
        table_name
            Nombre que tendra la tabla en el documento.

        sheet_name
            Nombre de la hoja donde se insertara la tabla.

        mode
            Modo de escritura
                `'w'`: Sobrescribe o crea un archivo nuevo.
                `'a'`: Agrega nuevas hojas al archivo existente.

        float_format
            Formato de los valores flotantes en el archivo `Excel` (ejemplo: `%.2f`).

        Lanza
        -----
        `FileNotFoundError`
            Si el documento no existe y se intenta usar el modo `'a'`.

        `ValueError`
            Si los nombres de la tabla, hoja o documento contienen caracteres no validos.

        Returns
        -------
        - Archivo `Excel` en binario dentro del buffer de la memoria.
        """
        try: 
            # Creamos un buffer para escribir el archivo Excel
            buffer = BytesIO()

            # Primero debemos almacenar el DataFrame en un fichero Excel
            with pd.ExcelWriter(buffer, mode=mode) as writer:
                self.to_excel(writer, sheet_name=sheet_name, index=False, float_format=float_format, na_rep=0)

            # Nos aseguramos de mover el puntero al inicio antes de leer
            buffer.seek(0)

            # Abrimos el fichero excel para poder trabajar sobre el 
            wb = openpyxl.load_workbook(filename=buffer)

            # Con el DataFrame podemos instanciar y dimensionar la tabla
            dimension = f'A1:{openpyxl.utils.get_column_letter(self.shape[1])}{len(self)+1}'
            table = openpyxl.worksheet.table.Table(displayName=table_name, ref=dimension)
            table.tableStyleInfo = STYLE_EXCEL

            # Una vez instanciada la tabla, la insertamos en el documento y ajustamos el ancho de sus columnas
            wb[sheet_name].add_table(table)
            wb = ProcessExcel.adjust_columns(wb, sheet_name)
            
            # Creamos un nuevo buffer final para la respuesta
            save_buffer = BytesIO()
            wb.save(save_buffer)
            save_buffer.seek(0)

            return save_buffer
        
        # Si hay un error, devolvemos None y manejamos la excepcion segun se necesite.
        except (ValueError, IOError) as e:
            return None

    @staticmethod
    def read_table(buffer_file: BytesIO, file_name: str, table: str) -> 'PdUtils':
        """
        Descripcion
        -----------
        Lee una tabla especifica de un archivo `Excel` y la convierte en un `DataFrame`.

        Parametros
        ----------
        buffer_file
            Archivo `Excel` en binario dentro del buffer de la memoria cargado a la aplicacion a traves de Flask.

        file_name
            Nombre del archivo `Excel` que se encuentra cargado en binario en la memoria.

        table
            Nombre de la tabla que se desea leer.

        Retorna
        -------
        - Contenido de la tabla en formato `DataFrame`.

        Lanza
        -----
        `FileNotFoundError`
            Si el archivo de `Excel` especificado no existe.

        `ValueError`
            Si la tabla especificada no se encuentra en el archivo.
        """
        # Abrimos el documento indicando las configuraciones de lectura que deseamos
        wb = openpyxl.load_workbook(filename=buffer_file, read_only=False, data_only=True)
        sheet_with_table = None
        tbl = None

        # Ahora vamos a iterar sobre cada hoja del libro
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]

            # Por cada hoja vamos a comprobrar si la tabla que buscamos existe dentro de esta hoja
            # Ya que pueden haber muchisimas tablas sobre una misma hoja
            # Y en el caso de obtenerla, extraeremos la tabla y almacenaremos la hoja, rompiendo posteriormente el bucle
            if table in sheet.tables:
                tbl = sheet.tables[table]
                sheet_with_table = sheet
                break
        
        # En caso de nunca conseguirla arrojaremos un error
        else:
            raise ValueError(f"La tabla `{table}` no se encontra en el documento `{file_name}`")
        
        # Sobre la tabla obtenida anteriormente, vamos a localizar sus dimensiones
        # Y sobre las dimensiones, los datos almacenamos
        tbl_range = tbl.ref
        data = sheet_with_table[tbl_range]

        # Iterando sobre los datos podremos obtener el valor en cada celda
        content = [[cell.value for cell in row] for row in data]

        # Del contenido de las celdas separamos el nombre de las columans del valor en cada celda de cada fila e instanciamos el DataFrame
        header = content[0]
        rest = content[1:]

        # Cerramos el libro de trabajo antes de retornar el dataframe
        wb.close()

        return PdUtils(rest, columns=header)
    
    @classmethod
    def validate_xlsx(self, buffer_file: BytesIO, file_name: str, table: str) -> tuple[bool, str]:
        """
        Descripcion
        -----------
        Realiza comprobaciones sobre el archivo `Excel` para verificar si cumple con los criterios para poder ser utilizado en la aplicacion.

        Parametros
        ----------
        buffer_file
            Archivo `Excel` en binario dentro del buffer de la memoria cargado a la aplicacion a traves de Flask.

        file_name
            Nombre del archivo `Excel` que se encuentra cargado en binario en la memoria.

        table
            Nombre de la tabla que se va a buscar dentro del documento `Excel`.

        Retorna 
        -------
        - `True` si el archivo es valido o `False` en caso contrario.
        - Mensaje relacionado con la validacion que finalizo la ejecucion del metodo.
        """
        # Vamos a instanciar un diccionario que nos permitira acceder de forma rapida a la verificacion de las columnas de la tabla
        dct_access = {
            'terciario': Templates.Terciario,
            'residencial': Templates.Residencial
        }

        # Controlaremos la lectura de los datos sobre el documento, ya que es lo primero a validar, que exista la tabla
        try:
            df = self.read_table(buffer_file, file_name, table)

            # Verificamos si la tabla contiene datos
            if df.empty:
                return False, f'La tabla `{table}`, en el documento `{file_name}`, no contiene datos 🔍'

            # Verificamos si todas las columnas existen
            if not set(Templates.obtain_keys_template(dct_access[table])).issubset(df.columns):
                return False, f'La tabla `{table}`, no contiene todas las columnas necesarias para el calculo, se recomienda usar la plantilla 🧐'

        # Si no se existe la tabla, lo indicaremos
        except ValueError:
            return False, f'La tabla `{table}` no existe dentro del documento `{file_name}` 🔍'
            
        # Si todas las comprobaciones son correctas, retornamos un simple true y un ok
        return True, 'Ok'
    
# -------------------------------------------------------------------------------------------------------------------------------------------------
# FIN DEL FICHERO :)
# -------------------------------------------------------------------------------------------------------------------------------------------------